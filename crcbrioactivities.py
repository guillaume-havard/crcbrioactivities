import openpyxl
import logging
import collections
import datetime

FILENAME = "tests/data/brio.xlsx"
INITIAL_SHEET_NAME = "Sheet1"

RECEIVED_100_PALETTES_RECEIVED_COLUMN_NAME = "Palettes sol 100*120 réellement reçues"
RECEIVED_80_PALETTES_RECEIVED_COLUMN_NAME = "Palettes sol 80*120 réellement reçues"
RECEIVED_80_EUR_PALETTES_RECEIVED_COLUMN_NAME = "Palettes EUR 80*120 réellement reçues"
STATUS_COLUMN_NAME = "tStatut"
LIVRAISON_REELLE_COLUMN_NAME = "TDate livraison réelle"
BL_COLUMN_NAME = "N° de BdL"

STATUS_OK = "livrée"
STATUS_NOT_OK = "annulée"

log = logging.getLogger(__name__)
formatter = logging.Formatter(
    "%(levelname)s %(asctime)s %(filename)s:%(lineno)d(%(funcName)s) %(message)s"
)
handler = logging.StreamHandler()
handler.setFormatter(formatter)
log.addHandler(handler)
log.setLevel(logging.DEBUG)


def values_column_id(row: int):
    """
    Generate an orderedDict {value: column id} for a row

    :param row: Row for which you want to generate the dict

    :return: The dict
    """
    columns_id = collections.OrderedDict()
    for row_data in sheet.iter_rows(min_row=row, max_row=row):
        columns_id = collections.OrderedDict(
            [(cell.value, id_column) for id_column, cell in enumerate(row_data)]
        )

    return columns_id


def analyse_nb_alettes_status(sheet, header_columns_id):
    """
    Analyse the sheet and return all rows where:
     * received palettes == 0 and status != STATUS_NOT_OK
     * received palettes != 0 and status == STATUS_NOT_OK

    :param sheet: Sheet with all information.
    :param header_columns_id: Dict of columns header with their id.

    :return: List of rows to analyse.
    """
    try:
        received_100_palettes_column = header_columns_id[RECEIVED_100_PALETTES_RECEIVED_COLUMN_NAME]
        received_80_palettes_column = header_columns_id[RECEIVED_80_PALETTES_RECEIVED_COLUMN_NAME]
        received_80_eur_palettes_column = header_columns_id[
            RECEIVED_80_EUR_PALETTES_RECEIVED_COLUMN_NAME
        ]
        status_column = header_columns_id[STATUS_COLUMN_NAME]
    except KeyError as error:
        log.error("The following column is nowhere to be found:", error)
        # TODO: stop function

    erroneous_rows = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        # TODO: test type of values
        if row[received_80_palettes_column].value:
            pal_80 = row[received_80_palettes_column].value
        else:
            pal_80 = 0
        if row[received_100_palettes_column].value:
            pal_100 = row[received_100_palettes_column].value
        else:
            pal_100 = 0
        if row[received_80_eur_palettes_column].value:
            pal_80_eur = row[received_80_eur_palettes_column].value
        else:
            pal_80_eur = 0

        total_received_palettes = pal_80 + pal_100 + pal_80_eur

        if total_received_palettes == 0 and row[status_column].value != STATUS_NOT_OK:
            erroneous_rows.append(row)
        elif total_received_palettes != 0 and row[status_column].value == STATUS_NOT_OK:
            erroneous_rows.append(row)

    return erroneous_rows


def add_sheet_nb_palettes_status(workbook, header_columns, rows):
    """
    Add a sheet to the workbook and write information about nb received palettes and status.

    :param workbook: Where the sheet will be added
    :param header_columns: Header of rows
    :param rows: Rows to write
    """
    palettes_worksheet = workbook.create_sheet("palettes")

    palettes_worksheet.append(["Analyse palettes"])
    palettes_worksheet.append([])
    palettes_worksheet.append(["Lignes avec incohérences :"])
    palettes_worksheet.append(list(header_columns.keys()))
    for row in rows:
        palettes_worksheet.append(row)


def analyse_date_livraison(sheet, header_columns_id):
    """
    TODO

    :param sheet: Sheet with all information.
    :param header_columns_id: Dict of columns header with their id.

    :return: List of rows to analyse.
    """
    try:
        livraison_reelle_column = header_columns_id[LIVRAISON_REELLE_COLUMN_NAME]
        status_column = header_columns_id[STATUS_COLUMN_NAME]
    except KeyError as error:
        log.error("The following column is nowhere to be found:", error)
        # TODO: stop function

    erroneous_rows = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        # TODO: test type of values

        date_str = row[livraison_reelle_column].value
        date_year = 0
        if date_str:
            if isinstance(date_str, str):
                date_year = int(date_str[len(date_str) - 4 :])
            elif isinstance(date_str, datetime.datetime):
                date_year = date_str.year

        if (not date_str or date_year < 2000) and row[status_column].value == STATUS_OK:
            erroneous_rows.append(row)

    return erroneous_rows


def add_sheet_date_status(workbook, header_columns, rows):
    """
    Add a sheet to the workbook and write information about TODO.

    :param workbook: Where the sheet will be added
    :param header_columns: Header of rows
    :param rows: Rows to write
    """
    palettes_worksheet = workbook.create_sheet("date")

    palettes_worksheet.append(["Analyse date livraison"])
    palettes_worksheet.append([])
    palettes_worksheet.append(["Lignes avec incohérences :"])
    palettes_worksheet.append(list(header_columns.keys()))
    for row in rows:
        palettes_worksheet.append(row)


def analyse_multiple_bl(sheet, header_columns_id):
    """
    TODO

    :param sheet: Sheet with all information.
    :param header_columns_id: Dict of columns header with their id.

    :return: List of rows to analyse.
    """
    try:
        bl_column = header_columns_id[BL_COLUMN_NAME]
    except KeyError as error:
        log.error("The following column is nowhere to be found:", error)
        # TODO: stop function

    bls = {}
    for id_row, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row)):
        # TODO: test type of values

        # TODO : multiple BL ?
        bls[str(row[bl_column].value)] = id_row

    # comptage racine BL
    bls_histo = {}
    for bl in bls:
        bl = str(bl)
        if bl.find("_") != -1:
            root_bl = bl[: bl.find("_")]
        else:
            root_bl = bl
        bl_histo_tmp = bls_histo.get(root_bl, [0, []])
        bl_histo_tmp[1].append(bl)
        bl_histo_tmp = [bl_histo_tmp[0] + 1, bl_histo_tmp[1]]

        bls_histo[root_bl] = bl_histo_tmp

    erroneous_rows = []
    for bl in bls_histo:
        if bls_histo[str(bl)][0] > 1:

            for bl_id in bls_histo[str(bl)][1]:
                # + 1 because we begin iwith ROW 1 in excel
                # +1 because we do not take into account the header
                erroneous_rows.append(sheet[bls[bl_id] + 1 + 1])

    return erroneous_rows


if __name__ == "__main__":
    file = openpyxl.load_workbook(FILENAME, data_only=True)
    print(file.sheetnames)
    sheet = file[INITIAL_SHEET_NAME]

    # Display sheet
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            print(cell.value, end=", ")
        print()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            print(cell.value, end=", ")
        print()

    header_columns_id = values_column_id(row=1)

    print()
    print(header_columns_id)

    palettes_erroneous_rows = analyse_nb_alettes_status(sheet, header_columns_id)

    print("erroneous_rows")
    for row in erroneous_rows:
        for cell in row:
            print(cell.value, end=", ")
        print()

    date_erroneous_rows = analyse_date_livraison(sheet, header_columns_id)
    # Creation of a workbook.
    output_workbook = openpyxl.Workbook(write_only=True)

    add_sheet_nb_palettes_status(output_workbook, header_columns_id, palettes_erroneous_rows)

    add_sheet_date_status(output_workbook, header_columns_id, date_erroneous_rows)
    # save workbook.
    output_workbook.save("analyse.xlsx")
